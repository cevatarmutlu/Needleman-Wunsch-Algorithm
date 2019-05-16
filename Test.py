import Fasta
import Needleman
import time
from multiprocessing import Process
from openpyxl import Workbook
import xlrd
import os


def isim_bulamadim(start: int, stop: int, line_list: list, excel_name):
    kitap = Workbook()
    excel = kitap.active
    excel['A1'] = "Seq1"
    excel['B1'] = "Seq2"
    excel['C1'] = "Skor"
    scores = []
    for i in range(start, stop):
        for j in range(i + 1, num_lines):
            score = Needleman.calculate(line_list[i], line_list[j], gap, match, mis_match)
            score_row = ['s' + str(i), 's' + str(j), str(score)]
            print(score_row)
            excel.append(score_row)
    kitap.save(excel_name)
    kitap.close()


line_list = Fasta.MyFasta('deneme').get_list()
line_list_arr = [line_list.copy() for x in range(3)]


match = 3.621354295
mis_match = -2.451795405
gap = -1.832482334
# num_lines = len(line_list)
# num_lines = 50
num_lines = 20


# Processes initialization
procs = []
p1 = Process(target=isim_bulamadim, args=(0, 3, line_list_arr[0], 'scores_val0.xlsx'))
p2 = Process(target=isim_bulamadim, args=(4, 11, line_list_arr[1], 'scores_val1.xlsx'))
p3 = Process(target=isim_bulamadim, args=(12, 20, line_list_arr[2], 'scores_val2.xlsx'))
# isim_bulamadim(0, 223, line_list1)
# isim_bulamadim(224, 454, line_list2)
# isim_bulamadim(455, 1000, line_list3)

procs.append(p1)
procs.append(p2)
procs.append(p3)

start_time = time.time()
print('Başlangıç zamanı %f' % start_time)
if __name__ == '__main__':
    for x in procs:
        x.start()
    for z in procs:
        z.join()

end_time = time.time()
print('Bitiş zamanı %f' % end_time)
bulma_suresi = end_time - start_time
print('Tamamlanma süresi %f' % (end_time - start_time))

wb = [xlrd.open_workbook('scores_val'+ str(x) +'.xlsx').sheet_by_index(0) for x in range(3)]


result = Workbook()
denn = result.active
denn['A1'] = "Seq1"
denn['B1'] = "Seq2"
denn['C1'] = "Skor"


for sheet in wb:
    for i in range(1, sheet.nrows):
        denn.append(sheet.row_values(i))

result.save('result.xlsx')
result.close()

os.remove('scores_val0.xlsx')
os.remove('scores_val1.xlsx')
os.remove('scores_val2.xlsx')


aa = xlrd.open_workbook('/home/cevat/PycharmProjects/Paralel/result.xlsx').sheet_by_index(0)
ff = []

for i in range(1, aa.nrows):
    ff.append(aa.row_values(i)[2])

zz = sorted(ff)[0:20]
zz_index = []

for z in zz:
    print(z)
    tt = ff.index(z)
    zz_index.append(tt)


result3 = Workbook()
denn3 = result3.active
denn3['A1'] = "Seq1"
denn3['B1'] = "Seq2"
denn3['C1'] = "Skor"
denn3['D1'] = str(bulma_suresi)

for gg in zz_index:
    print(gg, aa.row_values(gg + 1))
    denn3.append(aa.row_values(gg + 1))

result3.save('result3.xlsx')
result3.close()

os.remove('result.xlsx')
os.rename('result3.xlsx', 'result.xlsx')
print('Bitti :)')
